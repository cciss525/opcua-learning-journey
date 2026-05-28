"""生成 PLC OPC UA 支持情况对照表 (Excel)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "PLC OPC UA 支持对照"

# ── 样式 ──
header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
normal_font = Font(name="Microsoft YaHei", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── 标题 ──
ws.merge_cells("A1:F1")
ws["A1"] = "PLC 品牌型号 OPC UA 支持情况对照表"
ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14, color="2F5496")
ws["A1"].alignment = Alignment(horizontal="center")

# ── 表头 ──
headers = ["品牌", "系列/型号", "OPC UA 支持", "条件/说明", "国内常见度", "备注"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# ── 数据 ──
data = [
    # ── 原生支持 OPC UA ──
    ["西门子", "S7-1500 全系列", "是", "TIA Portal V14+ 启用 OPC UA Server 功能块", "★★★★★", "固件 ≥ 2.0 即支持"],
    ["西门子", "S7-1200（固件 ≥ 4.2）", "是", "需 TIA Portal 手动启用，仅 Server 功能", "★★★★★", "低固件版本不支持"],
    ["西门子", "S7-1200（固件 < 4.2）", "否", "需外接上位机或网关", "★★★★☆", "存量设备多"],
    ["西门子", "S7-300 / S7-400", "否", "S7 协议，需 Kepware/SIMATIC NET 转接", "★★★★★", "老产线主力"],
    ["西门子", "S7-200 / S7-200 SMART", "否", "PPI 协议，需网关", "★★★☆☆", "小型设备"],
    ["西门子", "ET 200SP", "是", "同 S7-1500 架构", "★★★★☆", "分布式 IO"],
    ["西门子", "LOGO!", "否", "微型 PLC，无 OPC UA", "★★★☆☆", "简单逻辑控制"],

    ["欧姆龙", "NJ/NX 系列（固件 ≥ 1.10）", "是", "Sysmac Studio 中配置 OPC UA Server", "★★★★☆", "中型主力"],
    ["欧姆龙", "NJ/NX 系列（固件 < 1.10）", "否", "需升级固件", "★★★☆☆", "老固件存量"],
    ["欧姆龙", "CJ 系列", "否", "FINS 协议，需 CJ2M + OPC UA 模块", "★★★★★", "量大，老产线"],
    ["欧姆龙", "CP 系列 (CP1H/CP1L)", "否", "Host Link / FINS，需上位机转换", "★★★★☆", "小型设备"],
    ["欧姆龙", "NX1P / NX102", "是", "内置 OPC UA Server", "★★★★☆", "新机型"],

    ["汇川", "AM600 系列（CODESYS 系）", "是*", "需购买 OPC UA 授权许可", "★★★★☆", "中型主力，授权费约几百元"],
    ["汇川", "AC800 系列（CODESYS 系）", "是*", "同 AM600，需付费授权", "★★★★☆", "大型控制器"],
    ["汇川", "H3U / H5U 系列", "否", "Modbus TCP / CANopen，无 OPC UA", "★★★★★", "国内出货量极大"],
    ["汇川", "H2U / H1U 系列", "否", "Modbus RTU/485", "★★★☆☆", "老型号，廉价"],
    ["汇川", "Easy 系列", "否", "Modbus TCP / 私有协议", "★★★★☆", "小型，性价比高"],

    ["三菱", "R 系列 (iQ-R)", "是", "需 OPC UA 模块 (RD81OPC96)", "★★★★☆", "高端，加模块不便宜"],
    ["三菱", "Q 系列 (iQ-F)", "否", "MC 协议 / SLMP，需网关", "★★★★★", "存量极大"],
    ["三菱", "FX 系列 (FX3U/FX5U)", "否", "MC 协议，需 MX OPC Server 或网关", "★★★★★", "全球出货量冠军级别"],
    ["三菱", "MELSEC iQ-F (FX5)", "是*", "部分固件支持，需额外配置", "★★★★☆", "较新"],

    ["罗克韦尔", "ControlLogix 5580 (L8x)", "是", "固件 ≥ 32，内嵌 OPC UA Server", "★★★☆☆", "美系高端"],
    ["罗克韦尔", "CompactLogix 5380 (5069)", "是", "固件 ≥ 32", "★★★☆☆", "美系中端"],
    ["罗克韦尔", "MicroLogix / SLC", "否", "EtherNet/IP / DF1，需 Kepware 转接", "★★☆☆☆", "老设备"],

    ["倍福", "TwinCAT 3（全系）", "是", "TF6100 OPC UA 功能，需购买授权", "★★★☆☆", "CODESYS 系，国内成长中"],
    ["倍福", "CX 系列嵌入式控制器", "是", "同 TwinCAT 3", "★★★☆☆", "紧凑型"],

    ["施耐德", "M340 / M580 (EcoStruxure)", "是", "固件升级后支持，需配置", "★★★☆☆", "过程控制"],
    ["施耐德", "M221 / M241", "否", "Modbus TCP，需 EcoStruxure OPC Server", "★★★☆☆", "小型"],

    ["台达", "AS 系列", "是*", "CODESYS 系，部分型号支持", "★★★★☆", "性价比高"],
    ["台达", "DVP 系列 (ES2/EX2)", "否", "Modbus RTU/TCP", "★★★★★", "国内量大"],
    ["台达", "AH 系列", "是*", "固件 ≥ 1.40", "★★★☆☆", "中大型"],

    ["ABB", "AC500 系列", "是", "Automation Builder 配置", "★★☆☆☆", "过程工业"],
    ["ABB", "AC500-eCo", "否", "需 OPC Server 软件", "★★☆☆☆", "经济型"],

    ["松下", "FP7 系列", "是", "需 OPC UA 对应固件", "★★☆☆☆", "日系高端"],
    ["松下", "FP-X / FP-XH", "否", "MEWTOCOL 协议", "★★★☆☆", "主流型号"],

    ["基恩士", "KV-8000 系列", "是", "内置 OPC UA Server", "★★☆☆☆", "高端视觉+PLC"],
    ["基恩士", "KV-5000/3000", "否", "上位链路协议", "★★★☆☆", "中端"],

    ["WAGO", "PFC200 (e!COCKPIT)", "是", "CODESYS 系，需授权", "★★☆☆☆", "楼宇/基础设施"],
]

# ── 写入数据 ──
for row_idx, row_data in enumerate(data, 4):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 第 3 列（OPC UA 支持）着色
    support = row_data[2]
    cell = ws.cell(row=row_idx, column=3)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if support == "是":
        cell.fill = yes_fill
    elif support.startswith("是*"):
        cell.fill = partial_fill
    else:
        cell.fill = no_fill

# ── 列宽 ──
widths = [10, 34, 10, 40, 12, 24]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = w

# ── 图例 sheet ──
ws2 = wb.create_sheet("图例说明")
ws2["A1"] = "图例说明"
ws2["A1"].font = Font(name="Microsoft YaHei", bold=True, size=12)
ws2["A3"] = "绿色 = 原生支持 OPC UA"
ws2["A3"].font = normal_font
ws2["A3"].fill = yes_fill
ws2["A4"] = "黄色 = 支持但需付费授权或条件限制"
ws2["A4"].font = normal_font
ws2["A4"].fill = partial_fill
ws2["A5"] = "红色 = 不支持 OPC UA（需外接网关/上位机）"
ws2["A5"].font = normal_font
ws2["A5"].fill = no_fill
ws2["A7"] = "国内常见度: ★ = 罕见  ★★★ = 常见  ★★★★★ = 市场主流"
ws2["A7"].font = normal_font
ws2["A9"] = f"生成日期: 2026-05-26 | 共 {len(data)} 个条目"
ws2["A9"].font = Font(name="Microsoft YaHei", size=9, italic=True, color="888888")
ws2.column_dimensions["A"].width = 60

# ── 冻结表头 + 筛选 ──
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:F{3 + len(data)}"

# ── 保存 ──
output = "D:/5ITlearning/OPCUA/data/PLC_OPC_UA_对照表.xlsx"
wb.save(output)
print(f"Done: {output}")
print(f"Total: {len(data)} entries (green={sum(1 for d in data if d[2]=='是')}, yellow={sum(1 for d in data if d[2].startswith('是*'))}, red={sum(1 for d in data if d[2]=='否')})")
